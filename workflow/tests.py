from io import BytesIO
from unittest.mock import patch

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse
from openpyxl import Workbook, load_workbook

from .forms import ProgressUpdateForm, TaskForm
from .models import ActivityLog, Assignment, Notification, Project, ProjectProgress, Task, TaskProgress, TelegramSettings
from .repositories import ProjectRepository, TaskRepository
from .services import ImportService, NotificationService, ProgressService, ProjectService, TaskService, build_projects_workbook

User = get_user_model()


class WorkflowServiceTests(TestCase):
    def setUp(self):
        self.manager = User.objects.create_user("manager", password="pass", role=User.Role.MANAGER)
        self.staff = User.objects.create_user("staff", password="pass", role=User.Role.STAFF, manager=self.manager)

    def test_assign_updates_current_employee_and_keeps_history(self):
        project = Project.objects.create(
            project_name="A",
            project_link="https://example.com/a",
            created_by=self.manager,
        )

        ProjectService.assign([project], self.staff, self.manager)
        project.refresh_from_db()

        self.assertEqual(project.current_employee, self.staff)
        self.assertEqual(project.status, Project.Status.ASSIGNED)
        self.assertEqual(Assignment.objects.count(), 1)
        self.assertEqual(ActivityLog.objects.filter(action=ActivityLog.Action.PROJECT_ASSIGNED).count(), 1)

    def test_assign_creates_deadline_priority_note_and_notification(self):
        from django.utils import timezone

        project = Project.objects.create(
            project_name="Deadline",
            project_link="https://example.com/deadline",
            created_by=self.manager,
        )
        deadline = timezone.now() + timezone.timedelta(days=1)

        ProjectService.assign(
            [project],
            self.staff,
            self.manager,
            deadline_at=deadline,
            priority=Project.Priority.URGENT,
            note="Lam gap",
            notify=True,
        )
        project.refresh_from_db()
        assignment = Assignment.objects.get(project=project)

        self.assertEqual(project.priority, Project.Priority.URGENT)
        self.assertEqual(project.deadline_at, deadline)
        self.assertEqual(assignment.note, "Lam gap")
        notice = Notification.objects.get(recipient=self.staff, notification_type=Notification.Type.PROJECT_ASSIGNED)
        self.assertIn(self.manager.username, notice.title)
        self.assertIn(project.project_name, notice.message)
        self.assertIn("Khẩn cấp", notice.message)

    def test_delete_project_removes_database_record(self):
        project = Project.objects.create(
            project_name="Delete Me",
            project_link="delete-me.example.com",
            created_by=self.manager,
        )

        count = ProjectService.soft_delete([project], self.manager)

        self.assertEqual(count, 1)
        self.assertFalse(Project.objects.filter(project_link="delete-me.example.com").exists())
        self.assertTrue(ActivityLog.objects.filter(action=ActivityLog.Action.PROJECT_DELETED).exists())

    def test_assign_sends_telegram_to_linked_staff(self):
        TelegramSettings.objects.create(enabled=True, bot_token="token")
        self.staff.telegram_enabled = True
        self.staff.telegram_chat_id = "12345"
        self.staff.save(update_fields=["telegram_enabled", "telegram_chat_id"])
        project = Project.objects.create(
            project_name="Telegram Project",
            project_link="https://example.com/telegram-project",
            created_by=self.manager,
        )

        with patch("workflow.services.TelegramService.send_message") as send_message:
            ProjectService.assign([project], self.staff, self.manager, notify=True)

        notice = Notification.objects.get(recipient=self.staff, notification_type=Notification.Type.PROJECT_ASSIGNED)
        self.assertEqual(notice.telegram_status, "SENT")
        send_message.assert_called_once()
        self.assertEqual(send_message.call_args.args[0], "12345")
        self.assertIn(project.project_name, send_message.call_args.args[1])

    def test_task_create_sends_telegram_to_linked_staff(self):
        TelegramSettings.objects.create(enabled=True, bot_token="token")
        self.staff.telegram_enabled = True
        self.staff.telegram_chat_id = "12345"
        self.staff.save(update_fields=["telegram_enabled", "telegram_chat_id"])
        form = TaskForm(
            data={
                "title": "Telegram Task",
                "description": "Do this task",
                "assignee": self.staff.pk,
                "priority": Task.Priority.HIGH,
                "status": Task.Status.NEW,
            }
        )
        self.assertTrue(form.is_valid(), form.errors)

        with patch("workflow.services.TelegramService.send_message") as send_message:
            task = TaskService.create_task(form, self.manager)

        notice = Notification.objects.get(recipient=self.staff, notification_type=Notification.Type.TASK_ASSIGNED)
        self.assertEqual(notice.telegram_status, "SENT")
        send_message.assert_called_once()
        self.assertEqual(send_message.call_args.args[0], "12345")
        self.assertIn(task.title, send_message.call_args.args[1])

    def test_project_progress_form_maps_stage_to_percent_and_note(self):
        expected = {
            "PENDING_REVIEW": (25, "Chờ Duyệt"),
            "CAMP_SET": (75, "Đã Set Camp"),
            "SPENT": (100, "Đã Chi Tiêu"),
        }
        for stage, (percent, note) in expected.items():
            with self.subTest(stage=stage):
                form = ProgressUpdateForm(data={"progress_stage": stage, "blocker_note": ""})

                self.assertTrue(form.is_valid(), form.errors)
                self.assertEqual(form.cleaned_data["progress_percent"], percent)
                self.assertEqual(form.cleaned_data["status_note"], note)

        form = ProgressUpdateForm(
            data={
                "progress_stage": "REGISTERED_SUCCESS",
                "registration_success_link": "success.example.com/path",
                "login_link": "login.example.com/path",
                "blocker_note": "",
            }
        )
        self.assertTrue(form.is_valid(), form.errors)
        self.assertEqual(form.cleaned_data["progress_percent"], 50)
        self.assertEqual(form.cleaned_data["status_note"], "ĐK Thành Công")
        self.assertEqual(form.cleaned_data["registration_success_link"], "https://success.example.com/path")
        self.assertEqual(form.cleaned_data["login_link"], "https://login.example.com/path")

    def test_project_progress_form_requires_link_for_registered_success(self):
        form = ProgressUpdateForm(data={"progress_stage": "REGISTERED_SUCCESS", "blocker_note": ""})

        self.assertFalse(form.is_valid())
        self.assertIn("registration_success_link", form.errors)
        self.assertIn("login_link", form.errors)

    def test_staff_progress_creates_update_and_manager_notification(self):
        project = Project.objects.create(
            project_name="Progress",
            project_link="https://example.com/progress",
            created_by=self.manager,
            current_employee=self.staff,
        )

        ProgressService.add_progress(project, self.staff, 45, "Dang xu ly", "Can tai khoan")

        self.assertEqual(ProjectProgress.objects.filter(project=project, progress_percent=45).count(), 1)
        notice = Notification.objects.get(notification_type=Notification.Type.PROGRESS_UPDATED)
        self.assertIn(self.staff.username, notice.title)
        self.assertIn(project.project_name, notice.message)
        self.assertIn("45%", notice.message)

    def test_registered_success_progress_saves_link_on_project(self):
        project = Project.objects.create(
            project_name="Registered Success",
            project_link="https://example.com/registered-success",
            created_by=self.manager,
            current_employee=self.staff,
        )

        ProgressService.add_progress(
            project,
            self.staff,
            50,
            "ĐK Thành Công",
            registration_success_link="https://success.example.com/account",
            login_link="https://success.example.com/login",
        )
        project.refresh_from_db()

        self.assertEqual(project.registration_success_link, "https://success.example.com/account")
        self.assertEqual(project.login_link, "https://success.example.com/login")

    def test_staff_status_update_notifies_manager(self):
        project = Project.objects.create(
            project_name="Status Update",
            project_link="https://example.com/status-update",
            created_by=self.manager,
            current_employee=self.staff,
            status=Project.Status.ASSIGNED,
        )

        ProjectService.update_status(project, Project.Status.WORKING, self.staff)

        notice = Notification.objects.get(notification_type=Notification.Type.STATUS_UPDATED)
        self.assertEqual(notice.recipient, self.manager)
        self.assertEqual(notice.actor, self.staff)
        self.assertIn(project.project_name, notice.message)
        self.assertFalse(Notification.objects.filter(recipient=self.staff, notification_type=Notification.Type.STATUS_UPDATED).exists())

    def test_staff_status_update_notifies_admin_and_direct_manager_only(self):
        admin = User.objects.create_user("admin-service", password="pass", role=User.Role.ADMIN, is_superuser=True)
        other_manager = User.objects.create_user("other-manager-service", password="pass", role=User.Role.MANAGER)
        project = Project.objects.create(
            project_name="Scoped Status Notice",
            project_link="https://example.com/scoped-status-notice",
            created_by=self.manager,
            manager=self.manager,
            current_employee=self.staff,
            status=Project.Status.ASSIGNED,
        )

        ProjectService.update_status(project, Project.Status.WORKING, self.staff)

        recipients = set(
            Notification.objects.filter(notification_type=Notification.Type.STATUS_UPDATED)
            .values_list("recipient__username", flat=True)
        )
        self.assertEqual(recipients, {admin.username, self.manager.username})
        self.assertNotIn(other_manager.username, recipients)

    def test_staff_progress_notifies_admin_and_direct_manager_only(self):
        admin = User.objects.create_user("admin-progress", password="pass", role=User.Role.ADMIN, is_superuser=True)
        other_manager = User.objects.create_user("other-manager-progress", password="pass", role=User.Role.MANAGER)
        project = Project.objects.create(
            project_name="Scoped Progress Notice",
            project_link="https://example.com/scoped-progress-notice",
            created_by=self.manager,
            manager=self.manager,
            current_employee=self.staff,
        )

        ProgressService.add_progress(project, self.staff, 50, "Dang lam")

        recipients = set(
            Notification.objects.filter(notification_type=Notification.Type.PROGRESS_UPDATED)
            .values_list("recipient__username", flat=True)
        )
        self.assertEqual(recipients, {admin.username, self.manager.username})
        self.assertNotIn(other_manager.username, recipients)

    def test_staff_status_update_sends_telegram_to_superuser_admin(self):
        TelegramSettings.objects.create(enabled=True, bot_token="token")
        admin = User.objects.create_user(
            "superadmin",
            password="pass",
            role=User.Role.STAFF,
            is_superuser=True,
            telegram_enabled=True,
            telegram_chat_id="99999",
        )
        project = Project.objects.create(
            project_name="Superuser Admin Notice",
            project_link="https://example.com/superuser-admin-notice",
            created_by=self.manager,
            current_employee=self.staff,
            status=Project.Status.ASSIGNED,
        )

        with patch("workflow.services.TelegramService.send_message") as send_message:
            ProjectService.update_status(project, Project.Status.WORKING, self.staff)

        notice = Notification.objects.get(
            recipient=admin,
            notification_type=Notification.Type.STATUS_UPDATED,
        )
        self.assertEqual(notice.telegram_status, "SENT")
        send_message.assert_called_once()
        self.assertEqual(send_message.call_args.args[0], "99999")
        self.assertIn(project.project_name, send_message.call_args.args[1])

    def test_task_progress_notification_identifies_actor_and_task(self):
        task = Task.objects.create(
            title="Soạn báo cáo tuần",
            description="Tổng hợp số liệu",
            assignee=self.staff,
            assigned_by=self.manager,
            priority=Task.Priority.HIGH,
        )

        TaskService.add_progress(task, self.staff, 60, "Đã tổng hợp dữ liệu", "Chờ duyệt số liệu")

        self.assertEqual(TaskProgress.objects.filter(task=task, progress_percent=60).count(), 1)
        notice = Notification.objects.get(notification_type=Notification.Type.TASK_PROGRESS_UPDATED)
        self.assertIn(self.staff.username, notice.title)
        self.assertIn(task.title, notice.message)
        self.assertIn("60%", notice.message)
        self.assertIn("Chờ duyệt số liệu", notice.message)

    def test_staff_task_progress_notifies_admin_and_direct_manager_only(self):
        admin = User.objects.create_user("admin-task-progress", password="pass", role=User.Role.ADMIN, is_superuser=True)
        other_manager = User.objects.create_user("other-manager-task-progress", password="pass", role=User.Role.MANAGER)
        task = Task.objects.create(
            title="Scoped Task Progress",
            description="Task scope",
            assignee=self.staff,
            assigned_by=self.manager,
        )

        TaskService.add_progress(task, self.staff, 80, "Gan xong")

        recipients = set(
            Notification.objects.filter(notification_type=Notification.Type.TASK_PROGRESS_UPDATED)
            .values_list("recipient__username", flat=True)
        )
        self.assertEqual(recipients, {admin.username, self.manager.username})
        self.assertNotIn(other_manager.username, recipients)

    def test_manager_only_sees_own_scope_tasks(self):
        other_manager = User.objects.create_user("other-task-manager", password="pass", role=User.Role.MANAGER)
        other_staff = User.objects.create_user("other-task-staff", password="pass", role=User.Role.STAFF, manager=other_manager)
        own_task = Task.objects.create(
            title="Own Task",
            description="Own",
            assignee=self.staff,
            assigned_by=self.manager,
        )
        delegated_team_task = Task.objects.create(
            title="Team Task",
            description="Team",
            assignee=self.staff,
            assigned_by=other_manager,
        )
        Task.objects.create(
            title="Hidden Task",
            description="Hidden",
            assignee=other_staff,
            assigned_by=other_manager,
        )

        visible = list(TaskRepository.visible_to(self.manager).order_by("title"))

        self.assertEqual(visible, [own_task, delegated_team_task])

    def test_staff_only_sees_assigned_projects(self):
        assigned = Project.objects.create(
            project_name="Assigned",
            project_link="https://example.com/assigned",
            created_by=self.manager,
            current_employee=self.staff,
        )
        Project.objects.create(
            project_name="Hidden",
            project_link="https://example.com/hidden",
            created_by=self.manager,
        )

        self.assertEqual(list(ProjectRepository.visible_to(self.staff)), [assigned])

    def test_manager_only_sees_own_scope_projects(self):
        other_manager = User.objects.create_user("other-manager", password="pass", role=User.Role.MANAGER)
        other_staff = User.objects.create_user("other-staff", password="pass", role=User.Role.STAFF, manager=other_manager)
        own_project = Project.objects.create(
            project_name="Own Manager Project",
            project_link="https://example.com/own-manager-project",
            created_by=self.manager,
            manager=self.manager,
        )
        team_project = Project.objects.create(
            project_name="Team Project",
            project_link="https://example.com/team-project",
            created_by=other_manager,
            current_employee=self.staff,
        )
        Project.objects.create(
            project_name="Hidden Other Manager Project",
            project_link="https://example.com/hidden-other-manager-project",
            created_by=other_manager,
            manager=other_manager,
            current_employee=other_staff,
        )

        visible = list(ProjectRepository.visible_to(self.manager).order_by("project_name"))

        self.assertEqual(visible, [own_project, team_project])

    def test_import_counts_database_and_file_duplicates(self):
        Project.objects.create(
            project_name="Existing",
            project_link="https://existing.example.com/path",
            created_by=self.manager,
        )
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Project Name", "Project Link"])
        sheet.append(["New", "https://new.example.com/path"])
        sheet.append(["Existing", "https://www.existing.example.com/other-path"])
        sheet.append(["New duplicate", "https://new.example.com/another-path"])
        sheet.append(["Missing link", ""])
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        buffer.name = "projects.xlsx"

        summary = ImportService.import_xlsx(buffer, self.manager)

        self.assertEqual(summary.batch.total_rows, 4)
        self.assertEqual(summary.batch.imported_rows, 1)
        self.assertEqual(summary.batch.duplicate_rows, 2)
        self.assertEqual(summary.batch.invalid_rows, 1)
        self.assertTrue(Project.objects.filter(project_link="new.example.com").exists())

    def test_import_accepts_link_only_and_generates_domain_project_name(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Link"])
        sheet.append(["https://example.com/my-new-project"])
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        buffer.name = "links.xlsx"

        summary = ImportService.import_xlsx(buffer, self.manager)
        project = summary.imported_projects[0]

        self.assertEqual(summary.batch.imported_rows, 1)
        self.assertEqual(project.project_name, "example.com")
        self.assertEqual(project.project_link, "example.com")

    def test_import_normalizes_domain_variants(self):
        summary = ImportService.import_pasted_links(
            "https://the5ers.com/\nhttps://WWW. the5ers.com/\n[https://the5ers.com/](https://the5ers.com/)",
            self.manager,
        )

        self.assertEqual(summary.batch.imported_rows, 1)
        self.assertEqual(summary.batch.duplicate_rows, 2)
        self.assertEqual(summary.imported_projects[0].project_link, "the5ers.com")

    def test_import_accepts_column_a_without_header(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["https://headerless.example.com/project"])
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        buffer.name = "column-a.xlsx"

        summary = ImportService.import_xlsx(buffer, self.manager)
        project = summary.imported_projects[0]

        self.assertEqual(summary.batch.imported_rows, 1)
        self.assertEqual(project.project_name, "headerless.example.com")

    def test_import_accepts_pasted_links(self):
        summary = ImportService.import_pasted_links(
            "https://manual-one.example.com/path\nhttps://manual-two.example.com/path",
            self.manager,
        )

        self.assertEqual(summary.batch.imported_rows, 2)
        self.assertTrue(Project.objects.filter(project_name="manual-one.example.com").exists())

    def test_import_treats_soft_deleted_project_link_as_duplicate(self):
        from django.utils import timezone

        Project.objects.create(
            project_name="Deleted",
            project_link="deleted.example.com",
            created_by=self.manager,
            deleted_at=timezone.now(),
        )

        summary = ImportService.import_pasted_links("https://deleted.example.com/path", self.manager)

        self.assertEqual(summary.batch.imported_rows, 0)
        self.assertEqual(summary.batch.duplicate_rows, 1)
        self.assertEqual(Project.objects.filter(project_link="deleted.example.com").count(), 1)

    def test_export_escapes_excel_formula_values(self):
        project = Project.objects.create(
            project_name="=HYPERLINK(\"https://example.com\")",
            project_link="https://example.com/formula",
            created_by=self.manager,
        )

        workbook = build_projects_workbook(Project.objects.filter(pk=project.pk))
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        exported = load_workbook(buffer, data_only=False)

        self.assertEqual(exported.active["A2"].value, "'=HYPERLINK(\"https://example.com\")")

    def test_import_rejects_too_many_rows(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Link"])
        for index in range(ImportService.MAX_ROWS + 1):
            sheet.append([f"https://example.com/{index}"])
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        buffer.name = "too-many.xlsx"

        with self.assertRaises(ValueError):
            ImportService.import_xlsx(buffer, self.manager)


class WorkflowViewTests(TestCase):
    def setUp(self):
        self.admin = User.objects.create_user(
            "admin",
            password="pass",
            role=User.Role.ADMIN,
            is_staff=True,
            is_superuser=True,
        )
        self.manager = User.objects.create_user("manager2", password="pass", role=User.Role.MANAGER)
        self.staff = User.objects.create_user("staff2", password="pass", role=User.Role.STAFF, manager=self.manager)
        self.project = Project.objects.create(
            project_name="View Project",
            project_link="https://example.com/view",
            created_by=self.manager,
            manager=self.manager,
        )

    def test_admin_can_open_user_management(self):
        self.client.login(username="admin", password="pass")
        response = self.client.get(reverse("user_list"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Tạo người dùng")

    def test_manager_can_assign_from_project_detail_action(self):
        self.client.login(username="manager2", password="pass")
        detail_response = self.client.get(self.project.get_absolute_url())

        self.assertContains(detail_response, 'name="notify" value="on"')

        response = self.client.post(
            reverse("assign_projects"),
            {"project_ids": str(self.project.pk), "employee": self.staff.pk, "notify": "on"},
        )
        self.project.refresh_from_db()

        self.assertEqual(response.status_code, 302)
        self.assertEqual(self.project.current_employee, self.staff)
        self.assertTrue(Notification.objects.filter(recipient=self.staff, project=self.project).exists())

    def test_manager_cannot_assign_project_to_staff_outside_team(self):
        other_manager = User.objects.create_user("outside-manager", password="pass", role=User.Role.MANAGER)
        outside_staff = User.objects.create_user("outside-staff", password="pass", role=User.Role.STAFF, manager=other_manager)
        self.client.login(username="manager2", password="pass")

        response = self.client.post(
            reverse("assign_projects"),
            {"project_ids": str(self.project.pk), "employee": outside_staff.pk, "notify": "on"},
        )
        self.project.refresh_from_db()

        self.assertEqual(response.status_code, 302)
        self.assertIsNone(self.project.current_employee)
        self.assertFalse(Notification.objects.filter(recipient=outside_staff, project=self.project).exists())

    def test_admin_can_bulk_assign_project_to_manager(self):
        project = Project.objects.create(
            project_name="Admin Assigned",
            project_link="https://example.com/admin-assigned",
            created_by=self.admin,
        )
        self.client.login(username="admin", password="pass")

        response = self.client.post(
            reverse("bulk_action"),
            {
                "action": "assign_manager",
                "project_ids": [str(project.pk)],
                "manager": str(self.manager.pk),
                "notify": "on",
            },
        )
        project.refresh_from_db()

        self.assertEqual(response.status_code, 302)
        self.assertEqual(project.manager, self.manager)
        self.assertTrue(Notification.objects.filter(recipient=self.manager, project=project).exists())

    def test_bulk_assign_accepts_deadline_priority_note_and_notification(self):
        from django.utils import timezone

        self.client.login(username="manager2", password="pass")
        deadline = timezone.now() + timezone.timedelta(days=2)
        response = self.client.post(
            reverse("bulk_action"),
            {
                "action": "assign",
                "project_ids": [str(self.project.pk)],
                "employee": str(self.staff.pk),
                "deadline_at": deadline.strftime("%Y-%m-%dT%H:%M"),
                "priority": Project.Priority.HIGH,
                "note": "Lam theo checklist",
                "notify": "on",
            },
        )
        self.project.refresh_from_db()
        assignment = Assignment.objects.get(project=self.project)

        self.assertEqual(response.status_code, 302)
        self.assertEqual(self.project.priority, Project.Priority.HIGH)
        self.assertEqual(assignment.note, "Lam theo checklist")
        self.assertTrue(Notification.objects.filter(recipient=self.staff, project=self.project).exists())

    def test_manager_can_bulk_change_project_state(self):
        self.client.login(username="manager2", password="pass")

        response = self.client.post(
            reverse("bulk_action"),
            {
                "action": "change_project_state",
                "project_ids": [str(self.project.pk)],
                "project_state": Project.ProjectState.PAUSED,
            },
        )
        self.project.refresh_from_db()

        self.assertEqual(response.status_code, 302)
        self.assertEqual(self.project.project_state, Project.ProjectState.PAUSED)

    def test_manager_can_update_project_state_from_detail(self):
        self.client.login(username="manager2", password="pass")

        response = self.client.post(
            reverse("project_state_update", args=[self.project.pk]),
            {"project_state": Project.ProjectState.AF_LOCKED},
        )
        self.project.refresh_from_db()

        self.assertEqual(response.status_code, 302)
        self.assertEqual(self.project.project_state, Project.ProjectState.AF_LOCKED)

    def test_manager_can_save_quick_project_update_once(self):
        self.client.login(username="manager2", password="pass")

        response = self.client.post(
            reverse("project_quick_update", args=[self.project.pk]),
            {
                "project_state": Project.ProjectState.PAUSED,
                "status": Project.Status.WORKING,
                "result": Project.Result.PROFIT,
            },
        )
        self.project.refresh_from_db()

        self.assertEqual(response.status_code, 302)
        self.assertEqual(self.project.project_state, Project.ProjectState.PAUSED)
        self.assertEqual(self.project.status, Project.Status.WORKING)
        self.assertEqual(self.project.result, Project.Result.PROFIT)

    def test_report_page_renders_chart_data(self):
        self.client.login(username="manager2", password="pass")
        response = self.client.get(reverse("reports"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "du-lieu-trang-thai")

    def test_manager_project_list_excludes_other_manager_projects(self):
        other_manager = User.objects.create_user("list-other-manager", password="pass", role=User.Role.MANAGER)
        other_staff = User.objects.create_user("list-other-staff", password="pass", role=User.Role.STAFF, manager=other_manager)
        hidden_project = Project.objects.create(
            project_name="Hidden Other Team",
            project_link="https://example.com/hidden-other-team",
            created_by=other_manager,
            manager=other_manager,
            current_employee=other_staff,
        )
        self.client.login(username="manager2", password="pass")

        response = self.client.get(reverse("project_list"))

        self.assertEqual(response.status_code, 200)
        projects = list(response.context["projects"])
        self.assertIn(self.project, projects)
        self.assertNotIn(hidden_project, projects)

    def test_manager_report_counts_only_own_scope_projects(self):
        other_manager = User.objects.create_user("report-other-manager", password="pass", role=User.Role.MANAGER)
        other_staff = User.objects.create_user("report-other-staff", password="pass", role=User.Role.STAFF, manager=other_manager)
        Project.objects.create(
            project_name="Hidden Report Project",
            project_link="https://example.com/hidden-report-project",
            created_by=other_manager,
            manager=other_manager,
            current_employee=other_staff,
        )
        self.client.login(username="manager2", password="pass")

        response = self.client.get(reverse("reports"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["counts"]["total"], 1)

    def test_manager_notification_list_excludes_other_team_notifications(self):
        other_manager = User.objects.create_user("notice-other-manager", password="pass", role=User.Role.MANAGER)
        other_staff = User.objects.create_user("notice-other-staff", password="pass", role=User.Role.STAFF, manager=other_manager)
        hidden_project = Project.objects.create(
            project_name="Hidden Notice Project",
            project_link="https://example.com/hidden-notice-project",
            created_by=other_manager,
            manager=other_manager,
            current_employee=other_staff,
        )
        visible_notice = Notification.objects.create(
            recipient=self.manager,
            actor=self.staff,
            project=self.project,
            title="Visible",
            message="Visible notice",
        )
        hidden_notice = Notification.objects.create(
            recipient=other_manager,
            actor=other_staff,
            project=hidden_project,
            title="Hidden",
            message="Hidden notice",
        )
        self.client.login(username="manager2", password="pass")

        response = self.client.get(reverse("notifications"))

        self.assertEqual(response.status_code, 200)
        notifications = list(response.context["notifications"])
        self.assertIn(visible_notice, notifications)
        self.assertNotIn(hidden_notice, notifications)

    def test_staff_can_submit_progress_for_assigned_project(self):
        self.project.current_employee = self.staff
        self.project.save(update_fields=["current_employee"])
        self.client.login(username="staff2", password="pass")

        response = self.client.post(
            reverse("project_progress_update", args=[self.project.pk]),
            {"progress_stage": "CAMP_SET", "blocker_note": ""},
        )

        self.assertEqual(response.status_code, 302)
        self.assertTrue(ProjectProgress.objects.filter(project=self.project, progress_percent=75, status_note="Đã Set Camp").exists())

    def test_staff_registered_success_progress_requires_and_saves_link(self):
        self.project.current_employee = self.staff
        self.project.save(update_fields=["current_employee"])
        self.client.login(username="staff2", password="pass")

        missing_response = self.client.post(
            reverse("project_progress_update", args=[self.project.pk]),
            {"progress_stage": "REGISTERED_SUCCESS", "blocker_note": ""},
        )
        self.project.refresh_from_db()
        self.assertEqual(missing_response.status_code, 302)
        self.assertEqual(self.project.registration_success_link, "")
        self.assertEqual(self.project.login_link, "")
        self.assertFalse(ProjectProgress.objects.filter(project=self.project, status_note="ĐK Thành Công").exists())

        response = self.client.post(
            reverse("project_progress_update", args=[self.project.pk]),
            {
                "progress_stage": "REGISTERED_SUCCESS",
                "registration_success_link": "success.example.com/dk",
                "login_link": "success.example.com/login",
                "blocker_note": "",
            },
        )
        self.project.refresh_from_db()

        self.assertEqual(response.status_code, 302)
        self.assertEqual(self.project.registration_success_link, "https://success.example.com/dk")
        self.assertEqual(self.project.login_link, "https://success.example.com/login")
        self.assertTrue(ProjectProgress.objects.filter(project=self.project, progress_percent=50, status_note="ĐK Thành Công").exists())

    def test_project_list_shows_registered_success_link_label(self):
        self.project.registration_success_link = "https://success.example.com/dk"
        self.project.login_link = "https://success.example.com/login"
        self.project.save(update_fields=["registration_success_link", "login_link"])
        self.client.login(username="manager2", password="pass")

        response = self.client.get(reverse("project_list"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'href="https://success.example.com/dk"')
        self.assertContains(response, 'href="https://success.example.com/login"')
        self.assertContains(response, "bi-link-45deg")

    def test_staff_can_update_assigned_project_state_and_result(self):
        self.project.current_employee = self.staff
        self.project.save(update_fields=["current_employee"])
        self.client.login(username="staff2", password="pass")

        state_response = self.client.post(
            reverse("project_state_update", args=[self.project.pk]),
            {"project_state": Project.ProjectState.PAUSED},
        )
        result_response = self.client.post(
            reverse("project_result_update", args=[self.project.pk]),
            {"result": Project.Result.PROFIT},
        )
        self.project.refresh_from_db()

        self.assertEqual(state_response.status_code, 302)
        self.assertEqual(result_response.status_code, 302)
        self.assertEqual(self.project.project_state, Project.ProjectState.PAUSED)
        self.assertEqual(self.project.result, Project.Result.PROFIT)

    def test_staff_cannot_update_unassigned_project_state_or_result(self):
        self.client.login(username="staff2", password="pass")

        state_response = self.client.post(
            reverse("project_state_update", args=[self.project.pk]),
            {"project_state": Project.ProjectState.PAUSED},
        )
        result_response = self.client.post(
            reverse("project_result_update", args=[self.project.pk]),
            {"result": Project.Result.PROFIT},
        )

        self.assertEqual(state_response.status_code, 404)
        self.assertEqual(result_response.status_code, 404)

    def test_staff_cannot_submit_progress_for_unassigned_project(self):
        self.client.login(username="staff2", password="pass")

        response = self.client.post(
            reverse("project_progress_update", args=[self.project.pk]),
            {"progress_stage": "CAMP_SET"},
        )

        self.assertEqual(response.status_code, 404)

    def test_notification_mark_read(self):
        notice = Notification.objects.create(
            recipient=self.staff,
            actor=self.manager,
            project=self.project,
            title="Test",
            message="Msg",
        )
        self.client.login(username="staff2", password="pass")

        response = self.client.post(reverse("notification_read", args=[notice.pk]))
        notice.refresh_from_db()

        self.assertEqual(response.status_code, 302)
        self.assertTrue(notice.is_read)

    def test_staff_activity_logs_only_show_related_entries(self):
        self.project.current_employee = self.staff
        self.project.save(update_fields=["current_employee"])
        hidden_project = Project.objects.create(
            project_name="Hidden Log Project",
            project_link="https://example.com/hidden-log",
            created_by=self.manager,
        )
        own_log = ActivityLog.objects.create(
            user=self.staff,
            project=None,
            action=ActivityLog.Action.PROGRESS_UPDATED,
            description="Own activity",
        )
        assigned_project_log = ActivityLog.objects.create(
            user=self.manager,
            project=self.project,
            action=ActivityLog.Action.PROJECT_ASSIGNED,
            description="Assigned project activity",
        )
        hidden_log = ActivityLog.objects.create(
            user=self.manager,
            project=hidden_project,
            action=ActivityLog.Action.PROJECT_UPDATED,
            description="Hidden activity",
        )
        self.client.login(username="staff2", password="pass")

        response = self.client.get(reverse("activity_logs"))

        self.assertEqual(response.status_code, 200)
        logs = list(response.context["logs"])
        self.assertIn(own_log, logs)
        self.assertIn(assigned_project_log, logs)
        self.assertNotIn(hidden_log, logs)

    def test_manager_activity_logs_show_all_entries(self):
        visible_log = ActivityLog.objects.create(
            user=self.staff,
            project=self.project,
            action=ActivityLog.Action.PROGRESS_UPDATED,
            description="Staff activity",
        )
        hidden_project = Project.objects.create(
            project_name="Manager Visible Project",
            project_link="https://example.com/manager-visible",
            created_by=self.manager,
        )
        other_log = ActivityLog.objects.create(
            user=self.manager,
            project=hidden_project,
            action=ActivityLog.Action.PROJECT_UPDATED,
            description="Manager activity",
        )
        self.client.login(username="manager2", password="pass")

        response = self.client.get(reverse("activity_logs"))

        self.assertEqual(response.status_code, 200)
        logs = list(response.context["logs"])
        self.assertIn(visible_log, logs)
        self.assertIn(other_log, logs)

    def test_telegram_settings_page_renders_instructions(self):
        TelegramSettings.objects.create(bot_username="demo_bot")
        self.client.login(username="staff2", password="pass")

        response = self.client.get(reverse("telegram_settings"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Hướng dẫn cho nhân viên")
        self.assertContains(response, "/start")

    def test_manager_can_update_general_ranking_setting(self):
        TelegramSettings.objects.create(show_employee_ranking_to_staff=True)
        self.client.login(username="manager2", password="pass")

        response = self.client.post(reverse("general_settings"), {"show_employee_ranking_to_staff": ""})

        self.assertEqual(response.status_code, 302)
        self.assertFalse(TelegramSettings.get_solo().show_employee_ranking_to_staff)

    def test_staff_cannot_open_general_settings(self):
        self.client.login(username="staff2", password="pass")

        response = self.client.get(reverse("general_settings"))

        self.assertEqual(response.status_code, 403)

    def test_notification_records_skipped_telegram_when_not_linked(self):
        notice = NotificationService.create(
            recipient=self.staff,
            actor=self.manager,
            project=self.project,
            title="Test Telegram",
            message="Msg",
        )

        self.assertEqual(notice.telegram_status, "SKIPPED")
        self.assertIn("chưa bật", notice.telegram_error)

    def test_telegram_notification_uses_custom_template(self):
        TelegramSettings.objects.create(
            enabled=True,
            bot_token="token",
            notification_template="Thông báo: {title}\nNgười gửi: {actor}\nDự án: {project}",
        )
        self.staff.telegram_enabled = True
        self.staff.telegram_chat_id = "12345"
        self.staff.save(update_fields=["telegram_enabled", "telegram_chat_id"])

        with patch("workflow.services.TelegramService.send_message") as send_message:
            NotificationService.create(
                recipient=self.staff,
                actor=self.manager,
                project=self.project,
                title="Cập nhật mới",
                message="Nội dung mới",
            )

        sent_text = send_message.call_args.args[1]
        self.assertIn("Thông báo: Cập nhật mới", sent_text)
        self.assertIn("Người gửi: manager2", sent_text)
        self.assertIn(f"Dự án: {self.project.project_name}", sent_text)

    def test_staff_dashboard_only_shows_visible_blockers(self):
        self.project.current_employee = self.staff
        self.project.save(update_fields=["current_employee"])
        hidden = Project.objects.create(
            project_name="Hidden blocker",
            project_link="https://example.com/hidden-blocker",
            created_by=self.manager,
        )
        ProjectProgress.objects.create(project=self.project, user=self.staff, progress_percent=10, status_note="Visible", blocker_note="visible blocker")
        ProjectProgress.objects.create(project=hidden, user=self.manager, progress_percent=20, status_note="Hidden", blocker_note="hidden blocker")
        self.client.login(username="staff2", password="pass")

        response = self.client.get(reverse("dashboard"))

        self.assertEqual(response.status_code, 200)
        blockers = list(response.context["blocked_progress"])
        self.assertEqual(len(blockers), 1)
        self.assertEqual(blockers[0].project, self.project)

    def test_staff_dashboard_hides_employee_ranking_when_disabled(self):
        TelegramSettings.objects.create(show_employee_ranking_to_staff=False)
        self.client.login(username="staff2", password="pass")

        response = self.client.get(reverse("dashboard"))

        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.context["show_employee_ranking"])
        self.assertNotContains(response, "Bảng xếp hạng nhân viên")

    def test_manager_dashboard_still_shows_employee_ranking_when_staff_setting_disabled(self):
        TelegramSettings.objects.create(show_employee_ranking_to_staff=False)
        self.client.login(username="manager2", password="pass")

        response = self.client.get(reverse("dashboard"))

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.context["show_employee_ranking"])
        self.assertContains(response, "Bảng xếp hạng nhân viên")
