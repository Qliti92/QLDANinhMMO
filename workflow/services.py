from dataclasses import dataclass
import json
import re
from urllib.error import HTTPError, URLError
from urllib.parse import urlparse
from urllib.request import Request, urlopen

from django.contrib.auth import get_user_model
from django.db import transaction
from django.db.models import Count, Q
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from .models import ActivityLog, Assignment, ImportBatch, Notification, Project, ProjectProgress, Task, TaskAttachment, TaskProgress, TelegramSettings

User = get_user_model()


def display_user(user) -> str:
    if not user:
        return "Hệ thống"
    full_name = getattr(user, "get_full_name", lambda: "")()
    return full_name or getattr(user, "username", str(user))


def display_datetime(value) -> str:
    if not value:
        return "chưa có hạn"
    return timezone.localtime(value).strftime("%d/%m/%Y %H:%M")


def display_choice(obj, field_name: str, fallback=None) -> str:
    getter = getattr(obj, f"get_{field_name}_display", None)
    if getter:
        return getter()
    return fallback if fallback is not None else getattr(obj, field_name, "")


def log_activity(user, action, description, project=None, metadata=None, request=None) -> ActivityLog:
    ip_address = None
    if request:
        forwarded = request.META.get("HTTP_X_FORWARDED_FOR")
        ip_address = forwarded.split(",")[0].strip() if forwarded else request.META.get("REMOTE_ADDR")
    return ActivityLog.objects.create(
        user=user if getattr(user, "is_authenticated", False) else None,
        project=project,
        action=action,
        description=description,
        metadata=metadata or {},
        ip_address=ip_address,
    )


@dataclass
class ImportSummary:
    batch: ImportBatch
    imported_projects: list[Project]


class ProjectService:
    @staticmethod
    @transaction.atomic
    def assign(projects, employee, assigned_by, request=None, deadline_at=None, priority=None, note="", notify=True) -> int:
        count = 0
        for project in projects:
            previous_employee = project.current_employee_id
            Assignment.objects.create(
                project=project,
                employee=employee,
                assigned_by=assigned_by,
                note=note or "",
                deadline_at=deadline_at,
                priority=priority or project.priority,
            )
            project.current_employee = employee
            project.status = Project.Status.ASSIGNED
            update_fields = ["current_employee", "status", "updated_at"]
            if deadline_at:
                project.deadline_at = deadline_at
                update_fields.append("deadline_at")
            if priority:
                project.priority = priority
                update_fields.append("priority")
            project.save(update_fields=update_fields)
            log_activity(
                assigned_by,
                ActivityLog.Action.PROJECT_ASSIGNED,
                f"Đã giao dự án cho {employee.username}",
                project=project,
                metadata={"previous_employee_id": previous_employee, "employee_id": employee.pk, "deadline_at": str(deadline_at or ""), "priority": priority or ""},
                request=request,
            )
            if notify:
                NotificationService.create(
                    recipient=employee,
                    actor=assigned_by,
                    project=project,
                    notification_type=Notification.Type.PROJECT_ASSIGNED,
                    title="",
                    message="",
                )
            count += 1
        return count

    @staticmethod
    @transaction.atomic
    def update_status(project, status, user, request=None) -> None:
        before = project.status
        project.status = status
        if status == Project.Status.DONE and not project.completed_at:
            project.completed_at = timezone.now()
        project.save(update_fields=["status", "completed_at", "updated_at"])
        log_activity(
            user,
            ActivityLog.Action.STATUS_CHANGED,
            f"Đổi trạng thái từ {before} sang {status}",
            project=project,
            metadata={"before": before, "after": status},
            request=request,
        )
        before_label = dict(Project.Status.choices).get(before, before)
        after_label = dict(Project.Status.choices).get(status, status)
        if project.current_employee and project.current_employee_id != getattr(user, "pk", None):
            NotificationService.create(
                recipient=project.current_employee,
                actor=user,
                project=project,
                notification_type=Notification.Type.STATUS_UPDATED,
                title=f"{display_user(user)} cập nhật trạng thái dự án",
                message=(
                    f"{display_user(user)} đã cập nhật trạng thái dự án “{project.project_name}” "
                    f"từ {before_label} sang {after_label}."
                ),
            )
        if not user.can_manage_projects:
            NotificationService.notify_managers(
                actor=user,
                project=project,
                title=f"{display_user(user)} cập nhật trạng thái dự án",
                message=(
                    f"{display_user(user)} đã cập nhật trạng thái dự án “{project.project_name}” "
                    f"từ {before_label} sang {after_label}."
                ),
                notification_type=Notification.Type.STATUS_UPDATED,
            )

    @staticmethod
    @transaction.atomic
    def update_result(project, result, user, request=None) -> None:
        before = project.result
        project.result = result
        project.result_updated_at = timezone.now()
        project.save(update_fields=["result", "result_updated_at", "updated_at"])
        log_activity(
            user,
            ActivityLog.Action.RESULT_UPDATED,
            f"Đổi kết quả từ {before} sang {result}",
            project=project,
            metadata={"before": before, "after": result},
            request=request,
        )
        if project.current_employee and project.current_employee_id != getattr(user, "pk", None):
            NotificationService.create(
                recipient=project.current_employee,
                actor=user,
                project=project,
                notification_type=Notification.Type.RESULT_UPDATED,
                title=f"{display_user(user)} cập nhật kết quả dự án",
                message=(
                    f"{display_user(user)} đã cập nhật kết quả dự án “{project.project_name}” "
                    f"từ {dict(Project.Result.choices).get(before, before)} sang {dict(Project.Result.choices).get(result, result)}."
                ),
            )
        if not user.can_manage_projects:
            before_label = dict(Project.Result.choices).get(before, before)
            after_label = dict(Project.Result.choices).get(result, result)
            NotificationService.notify_managers(
                actor=user,
                project=project,
                title=f"{display_user(user)} cập nhật kết quả dự án",
                message=(
                    f"{display_user(user)} đã cập nhật kết quả dự án “{project.project_name}” "
                    f"từ {before_label} sang {after_label}."
                ),
                notification_type=Notification.Type.RESULT_UPDATED,
            )

    @staticmethod
    @transaction.atomic
    def soft_delete(projects, user, request=None) -> int:
        count = 0
        for project in projects:
            project.soft_delete(user=user)
            log_activity(
                user,
                ActivityLog.Action.PROJECT_DELETED,
                "Đã xóa mềm dự án",
                project=project,
                request=request,
            )
            count += 1
        return count


class NotificationService:
    @staticmethod
    def create(recipient, title, message, notification_type=Notification.Type.SYSTEM, actor=None, project=None, task=None):
        if notification_type == Notification.Type.PROJECT_ASSIGNED and project:
            actor_name = display_user(actor)
            title = f"{actor_name} giao dự án cho bạn"
            message = (
                f"{actor_name} đã giao dự án “{project.project_name}” cho bạn. "
                f"Hạn xử lý: {display_datetime(project.deadline_at)}. "
                f"Độ ưu tiên: {display_choice(project, 'priority')}."
            )
        elif notification_type == Notification.Type.TASK_ASSIGNED and task:
            actor_name = display_user(actor)
            title = f"{actor_name} giao nhiệm vụ cho bạn"
            message = (
                f"{actor_name} đã giao nhiệm vụ “{task.title}” cho bạn. "
                f"Hạn hoàn thành: {display_datetime(task.deadline_at)}. "
                f"Độ ưu tiên: {display_choice(task, 'priority')}."
            )
        notification = Notification.objects.create(
            recipient=recipient,
            actor=actor,
            project=project,
            task=task,
            title=title,
            message=message,
            notification_type=notification_type,
        )
        TelegramService.send_notification(notification)
        return notification

    @staticmethod
    def notify_managers(actor, project, title, message, notification_type, task=None):
        recipients = User.objects.filter(
            Q(role__in=[User.Role.ADMIN, User.Role.MANAGER]) | Q(is_superuser=True),
            is_active=True,
        ).exclude(pk=getattr(actor, "pk", None))
        for recipient in recipients:
            NotificationService.create(recipient, title, message, notification_type, actor=actor, project=project, task=task)


class ProgressService:
    @staticmethod
    @transaction.atomic
    def add_progress(project, user, progress_percent, status_note, blocker_note="", request=None):
        progress = ProjectProgress.objects.create(
            project=project,
            user=user,
            progress_percent=progress_percent,
            status_note=status_note,
            blocker_note=blocker_note or "",
        )
        log_activity(
            user,
            ActivityLog.Action.PROGRESS_UPDATED,
            f"Cập nhật tiến trình {progress_percent}%",
            project=project,
            metadata={"progress_percent": progress_percent, "blocker_note": blocker_note or ""},
            request=request,
        )
        NotificationService.notify_managers(
            actor=user,
            project=project,
            title=f"{display_user(user)} thêm tiến độ dự án",
            message=(
                f"{display_user(user)} đã thêm tiến độ {progress_percent}% cho dự án "
                f"“{project.project_name}”. Nội dung: {status_note}"
                + (f" Vướng mắc: {blocker_note}" if blocker_note else "")
            ),
            notification_type=Notification.Type.PROGRESS_UPDATED,
        )
        return progress


class TelegramService:
    API_BASE = "https://api.telegram.org/bot{token}/{method}"

    @classmethod
    def is_ready_for(cls, user) -> bool:
        settings = TelegramSettings.get_solo()
        return bool(
            settings.enabled
            and settings.bot_token
            and user
            and user.telegram_enabled
            and user.telegram_chat_id
        )

    @classmethod
    def api_call(cls, method, payload=None):
        settings = TelegramSettings.get_solo()
        if not settings.bot_token:
            raise ValueError("Chưa cấu hình bot token Telegram.")
        url = cls.API_BASE.format(token=settings.bot_token, method=method)
        data = json.dumps(payload or {}).encode("utf-8")
        request = Request(url, data=data, headers={"Content-Type": "application/json"}, method="POST")
        try:
            with urlopen(request, timeout=8) as response:
                body = response.read().decode("utf-8")
        except HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="replace")
            raise ValueError(f"Telegram trả lỗi HTTP {exc.code}: {detail}") from exc
        except URLError as exc:
            raise ValueError(f"Không kết nối được Telegram: {exc.reason}") from exc
        result = json.loads(body)
        if not result.get("ok"):
            raise ValueError(result.get("description") or "Telegram API trả về lỗi.")
        return result.get("result")

    @classmethod
    def send_message(cls, chat_id, text):
        return cls.api_call(
            "sendMessage",
            {
                "chat_id": chat_id,
                "text": text,
                "disable_web_page_preview": True,
            },
        )

    @classmethod
    def format_notification(cls, notification) -> str:
        object_line = ""
        if notification.task:
            object_line = f"\n\nNhiệm vụ: {notification.task.title}"
        elif notification.project:
            object_line = f"\n\nDự án: {notification.project.project_name}"
        values = {
            "title": notification.title,
            "message": notification.message,
            "actor": display_user(notification.actor),
            "recipient": display_user(notification.recipient),
            "project": notification.project.project_name if notification.project else "",
            "task": notification.task.title if notification.task else "",
            "type": notification.notification_type,
            "object_line": object_line,
        }
        settings = TelegramSettings.get_solo()
        template = settings.notification_template or TelegramSettings.DEFAULT_NOTIFICATION_TEMPLATE
        try:
            return template.format(**values).strip()
        except (KeyError, ValueError):
            return TelegramSettings.DEFAULT_NOTIFICATION_TEMPLATE.format(**values).strip()
        lines = [
            f"🔔 {notification.title}",
            "",
            notification.message,
        ]
        if notification.task:
            lines.extend(["", f"Nhiệm vụ: {notification.task.title}"])
        elif notification.project:
            lines.extend(["", f"Dự án: {notification.project.project_name}"])
        return "\n".join(lines)

    @classmethod
    def send_notification(cls, notification):
        user = notification.recipient
        if not cls.is_ready_for(user):
            notification.telegram_status = "SKIPPED"
            notification.telegram_error = "Telegram chưa bật hoặc tài khoản chưa liên kết chat ID."
            notification.save(update_fields=["telegram_status", "telegram_error"])
            return False
        try:
            cls.send_message(user.telegram_chat_id, cls.format_notification(notification))
        except ValueError as exc:
            notification.telegram_status = "FAILED"
            notification.telegram_error = str(exc)
            notification.save(update_fields=["telegram_status", "telegram_error"])
            return False
        notification.telegram_status = "SENT"
        notification.telegram_error = ""
        notification.telegram_sent_at = timezone.now()
        notification.save(update_fields=["telegram_status", "telegram_error", "telegram_sent_at"])
        return True

    @classmethod
    def sync_updates(cls):
        settings = TelegramSettings.get_solo()
        offset = (settings.last_update_id + 1) if settings.last_update_id else None
        payload = {"timeout": 1, "allowed_updates": ["message"]}
        if offset:
            payload["offset"] = offset
        linked = []
        try:
            updates = cls.api_call("getUpdates", payload)
            for update in updates:
                update_id = update.get("update_id")
                if update_id is not None:
                    settings.last_update_id = max(settings.last_update_id or 0, update_id)
                message = update.get("message") or {}
                text = (message.get("text") or "").strip()
                if not text.startswith("/start"):
                    continue
                parts = text.split(maxsplit=1)
                if len(parts) < 2:
                    continue
                token = parts[1].strip()
                chat = message.get("chat") or {}
                user = User.objects.filter(telegram_link_token=token).first()
                if not user:
                    continue
                user.telegram_chat_id = str(chat.get("id", ""))
                user.telegram_username = chat.get("username", "") or ""
                user.telegram_enabled = True
                user.save(update_fields=["telegram_chat_id", "telegram_username", "telegram_enabled"])
                linked.append(user)
            settings.last_error = ""
        except ValueError as exc:
            settings.last_error = str(exc)
        settings.last_sync_at = timezone.now()
        settings.save(update_fields=["last_update_id", "last_error", "last_sync_at", "updated_at"])
        return linked

    @classmethod
    def test_bot(cls):
        result = cls.api_call("getMe", {})
        settings = TelegramSettings.get_solo()
        username = result.get("username", "")
        if username and not settings.bot_username:
            settings.bot_username = username
            settings.save(update_fields=["bot_username", "updated_at"])
        return result


class TaskService:
    @staticmethod
    @transaction.atomic
    def create_task(form, assigned_by, files=None, request=None):
        task = form.save(commit=False)
        task.assigned_by = assigned_by
        task.save()
        for uploaded in files or []:
            TaskAttachment.objects.create(
                task=task,
                file=uploaded,
                original_filename=uploaded.name,
                uploaded_by=assigned_by,
            )
        assigned_by_name = display_user(assigned_by)
        NotificationService.create(
            recipient=task.assignee,
            actor=assigned_by,
            task=task,
            notification_type=Notification.Type.TASK_ASSIGNED,
            title=f"{assigned_by_name} giao nhiệm vụ cho bạn",
            message=(
                f"{assigned_by_name} đã giao nhiệm vụ “{task.title}” cho bạn. "
                f"Hạn hoàn thành: {display_datetime(task.deadline_at)}. "
                f"Độ ưu tiên: {task.get_priority_display()}. "
                f"File đính kèm: {len(files or [])}."
            ),
        )
        log_activity(
            assigned_by,
            ActivityLog.Action.TASK_CREATED,
            f"Đã giao nhiệm vụ {task.title} cho {task.assignee.username}",
            metadata={"task_id": task.pk, "assignee_id": task.assignee_id},
            request=request,
        )
        return task

    @staticmethod
    @transaction.atomic
    def update_task(task, form, user, files=None, request=None):
        previous_assignee_id = task.assignee_id
        task = form.save()
        uploaded_count = len(files or [])
        for uploaded in files or []:
            TaskAttachment.objects.create(
                task=task,
                file=uploaded,
                original_filename=uploaded.name,
                uploaded_by=user,
            )
        log_activity(
            user,
            ActivityLog.Action.TASK_UPDATED,
            f"Đã cập nhật nhiệm vụ {task.title}",
            metadata={"task_id": task.pk},
            request=request,
        )
        if previous_assignee_id != task.assignee_id:
            NotificationService.create(
                recipient=task.assignee,
                actor=user,
                task=task,
                notification_type=Notification.Type.TASK_ASSIGNED,
                title="",
                message="",
            )
        elif task.assignee_id != getattr(user, "pk", None):
            NotificationService.create(
                recipient=task.assignee,
                actor=user,
                task=task,
                notification_type=Notification.Type.STATUS_UPDATED,
                title=f"{display_user(user)} cập nhật nhiệm vụ của bạn",
                message=(
                    f"{display_user(user)} đã cập nhật nhiệm vụ “{task.title}”. "
                    f"Trạng thái hiện tại: {task.get_status_display()}. "
                    f"Hạn hoàn thành: {display_datetime(task.deadline_at)}. "
                    f"File mới: {uploaded_count}."
                ),
            )
        return task

    @staticmethod
    @transaction.atomic
    def update_status(task, status, user, request=None):
        before = task.status
        task.status = status
        if status == Task.Status.DONE and not task.completed_at:
            task.completed_at = timezone.now()
        task.save(update_fields=["status", "completed_at", "updated_at"])
        log_activity(
            user,
            ActivityLog.Action.TASK_UPDATED,
            f"Cập nhật trạng thái nhiệm vụ từ {before} sang {status}",
            metadata={"task_id": task.pk, "before": before, "after": status},
            request=request,
        )
        before_label = dict(Task.Status.choices).get(before, before)
        after_label = dict(Task.Status.choices).get(status, status)
        if user.pk == task.assignee_id:
            NotificationService.notify_managers(
                actor=user,
                project=None,
                task=task,
                title=f"{display_user(user)} cập nhật trạng thái nhiệm vụ",
                message=(
                    f"{display_user(user)} đã cập nhật trạng thái nhiệm vụ “{task.title}” "
                    f"từ {before_label} sang {after_label}."
                ),
                notification_type=Notification.Type.TASK_PROGRESS_UPDATED,
            )
        elif task.assignee_id != user.pk:
            NotificationService.create(
                recipient=task.assignee,
                actor=user,
                task=task,
                notification_type=Notification.Type.STATUS_UPDATED,
                title=f"{display_user(user)} cập nhật nhiệm vụ của bạn",
                message=(
                    f"{display_user(user)} đã cập nhật trạng thái nhiệm vụ “{task.title}” "
                    f"từ {before_label} sang {after_label}."
                ),
            )

    @staticmethod
    @transaction.atomic
    def add_progress(task, user, progress_percent, status_note, blocker_note="", request=None):
        progress = TaskProgress.objects.create(
            task=task,
            user=user,
            progress_percent=progress_percent,
            status_note=status_note,
            blocker_note=blocker_note or "",
        )
        log_activity(
            user,
            ActivityLog.Action.TASK_PROGRESS_UPDATED,
            f"Cập nhật tiến độ nhiệm vụ {progress_percent}%",
            metadata={"task_id": task.pk, "progress_percent": progress_percent, "blocker_note": blocker_note or ""},
            request=request,
        )
        NotificationService.notify_managers(
            actor=user,
            project=None,
            task=task,
            title=f"{display_user(user)} thêm tiến độ nhiệm vụ",
            message=(
                f"{display_user(user)} đã thêm tiến độ {progress_percent}% cho nhiệm vụ "
                f"“{task.title}”. Nội dung: {status_note}"
                + (f" Vướng mắc: {blocker_note}" if blocker_note else "")
            ),
            notification_type=Notification.Type.TASK_PROGRESS_UPDATED,
        )
        return progress


class ImportService:
    MAX_ROWS = 10000
    NAME_COLUMNS = {"Project Name", "Tên dự án", "Ten du an"}
    LINK_COLUMNS = {
        "Link",
        "Project Link",
        "Link dự án",
        "Link du an",
        "Liên kết dự án",
        "Lien ket du an",
        "URL",
    }

    @classmethod
    @transaction.atomic
    def import_xlsx(cls, uploaded_file, user, request=None) -> ImportSummary:
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
        sheet = workbook.active
        if sheet.max_row and sheet.max_row > cls.MAX_ROWS + 1:
            raise ValueError(f"Excel import is limited to {cls.MAX_ROWS} data rows.")
        rows = list(sheet.iter_rows(values_only=True))
        headers = [str(value).strip() if value is not None else "" for value in (rows[0] if rows else [])]
        name_idx = cls._find_header(headers, cls.NAME_COLUMNS)
        link_idx = cls._find_header(headers, cls.LINK_COLUMNS)
        if link_idx is None:
            link_idx = 0
            data_rows = enumerate(rows, start=1)
        else:
            data_rows = enumerate(rows[1:], start=2)
        batch = ImportBatch.objects.create(uploaded_by=user, original_filename=uploaded_file.name)
        return cls._import_rows(data_rows, name_idx, link_idx, batch, user, request=request)

    @classmethod
    @transaction.atomic
    def import_pasted_links(cls, links_text, user, request=None) -> ImportSummary:
        rows = [("", line.strip()) for line in links_text.splitlines() if line.strip()]
        batch = ImportBatch.objects.create(uploaded_by=user, original_filename="nhap-link-thu-cong.txt")
        return cls._import_rows(enumerate(rows, start=1), 0, 1, batch, user, request=request)

    @classmethod
    def _import_rows(cls, data_rows, name_idx, link_idx, batch, user, request=None) -> ImportSummary:
        seen_domains = set()
        existing_domains = {
            cls.normalize_domain(project.project_link): project
            for project in Project.objects.select_related("current_employee")
        }
        existing_domains.pop("", None)
        imported = []
        duplicate_report = []
        invalid_report = []

        for number, row in data_rows:
            name = (
                str(row[name_idx]).strip()
                if name_idx is not None and len(row) > name_idx and row[name_idx]
                else ""
            )
            link = str(row[link_idx]).strip() if len(row) > link_idx and row[link_idx] else ""
            if not name and not link:
                continue
            batch.total_rows += 1
            if not link:
                invalid_report.append({"row": number, "project_name": name, "project_link": link, "reason": "Thiếu link"})
                batch.invalid_rows += 1
                continue
            domain = cls.normalize_domain(link)
            if not domain:
                invalid_report.append({"row": number, "project_name": name, "project_link": link, "reason": "Không lấy được domain"})
                batch.invalid_rows += 1
                continue
            if not name:
                name = cls.project_name_from_link(domain)
            if domain in seen_domains:
                duplicate_report.append({"row": number, "project_name": name, "project_link": domain, "source": "file"})
                batch.duplicate_rows += 1
                continue
            seen_domains.add(domain)

            existing = existing_domains.get(domain)
            if existing:
                duplicate_report.append(
                    {
                        "row": number,
                        "project_name": name,
                        "project_link": domain,
                        "source": "database",
                        "current_assignee": existing.current_employee.username if existing.current_employee else "",
                        "current_status": existing.status,
                        "current_result": existing.result,
                    }
                )
                batch.duplicate_rows += 1
                continue

            project = Project.objects.create(
                project_name=name,
                project_link=domain,
                created_by=user,
                import_batch=batch,
            )
            existing_domains[domain] = project
            imported.append(project)
            batch.imported_rows += 1
            log_activity(
                user,
                ActivityLog.Action.PROJECT_IMPORTED,
                "Nhập dự án từ Excel",
                project=project,
                metadata={"batch_id": batch.pk},
                request=request,
            )

        batch.duplicate_report = duplicate_report
        batch.invalid_report = invalid_report
        batch.save()
        return ImportSummary(batch=batch, imported_projects=imported)

    @staticmethod
    def _find_header(headers, candidates):
        for candidate in candidates:
            if candidate in headers:
                return headers.index(candidate)
        return None

    @classmethod
    def normalize_domain(cls, value: str) -> str:
        raw = str(value or "").strip()
        if not raw:
            return ""
        raw = raw.replace("\u00a0", " ")
        raw = re.sub(r"\.\s+", ".", raw)
        raw = re.sub(r"\s+\.", ".", raw)
        match = re.search(r"https?://[^\]\)\s]+", raw, flags=re.IGNORECASE)
        candidate = match.group(0) if match else raw
        candidate = candidate.strip("[]()<>\"'` \t\r\n")
        if "://" not in candidate:
            candidate = f"https://{candidate}"
        parsed = urlparse(candidate)
        domain = parsed.netloc or parsed.path.split("/")[0]
        domain = domain.split("@")[-1].split(":")[0].strip().lower()
        if domain.startswith("www."):
            domain = domain[4:]
        if "." not in domain:
            return ""
        return domain

    @classmethod
    def project_name_from_link(cls, link: str) -> str:
        return cls.normalize_domain(link)[:255]


class ReportService:
    @staticmethod
    def dashboard_counts(qs):
        now = timezone.now()
        active = qs.exclude(status__in=[Project.Status.DONE, Project.Status.CANCELLED])
        progress_values = [
            latest.progress_percent
            for project in qs.prefetch_related("progress_updates")
            for latest in [project.latest_progress]
            if latest
        ]
        return {
            "total": qs.count(),
            "new": qs.filter(status=Project.Status.NEW).count(),
            "assigned": qs.filter(status=Project.Status.ASSIGNED).count(),
            "working": qs.filter(status=Project.Status.WORKING).count(),
            "completed": qs.filter(status=Project.Status.DONE).count(),
            "cancelled": qs.filter(status=Project.Status.CANCELLED).count(),
            "profit": qs.filter(result=Project.Result.PROFIT).count(),
            "loss": qs.filter(result=Project.Result.LOSS).count(),
            "pending_result": qs.filter(result=Project.Result.PENDING).count(),
            "overdue": active.filter(deadline_at__lt=now).count(),
            "due_soon": active.filter(deadline_at__gte=now, deadline_at__lte=now + timezone.timedelta(hours=24)).count(),
            "no_deadline": active.filter(deadline_at__isnull=True).count(),
            "high_priority": qs.filter(priority=Project.Priority.HIGH).count(),
            "urgent_priority": qs.filter(priority=Project.Priority.URGENT).count(),
            "updated_today": qs.filter(updated_at__date=timezone.localdate()).count(),
            "avg_progress": round(sum(progress_values) / len(progress_values), 1) if progress_values else 0,
        }

    @staticmethod
    def employee_kpis():
        users = User.objects.filter(role=User.Role.STAFF).annotate(
            total_assigned=Count("current_projects", filter=Q(current_projects__deleted_at__isnull=True)),
            completed=Count("current_projects", filter=Q(current_projects__status=Project.Status.DONE, current_projects__deleted_at__isnull=True)),
            profit=Count("current_projects", filter=Q(current_projects__result=Project.Result.PROFIT, current_projects__deleted_at__isnull=True)),
            loss=Count("current_projects", filter=Q(current_projects__result=Project.Result.LOSS, current_projects__deleted_at__isnull=True)),
            overdue=Count("current_projects", filter=Q(current_projects__deadline_at__lt=timezone.now(), current_projects__deleted_at__isnull=True)),
        )
        rows = []
        for user in users:
            success_rate = (user.profit / user.completed * 100) if user.completed else 0
            progress_values = [item.progress_percent for item in user.progress_updates.all()]
            avg_progress = round(sum(progress_values) / len(progress_values), 1) if progress_values else 0
            rows.append({"user": user, "success_rate": round(success_rate, 2), "avg_progress": avg_progress})
        return sorted(rows, key=lambda row: (row["success_rate"], row["user"].profit, row["avg_progress"]), reverse=True)


def build_projects_workbook(projects):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Du an"
    sheet.append(["Tên dự án", "Domain", "Nhân viên", "Trạng thái dự án", "Trạng thái công việc", "Kết quả", "Ngày tạo"])
    for project in projects:
        sheet.append(
            [
                excel_safe(project.project_name),
                excel_safe(project.project_link),
                excel_safe(project.current_employee.username if project.current_employee else ""),
                excel_safe(project.get_project_state_display()),
                excel_safe(project.get_status_display()),
                excel_safe(project.get_result_display()),
                project.created_at.strftime("%Y-%m-%d %H:%M"),
            ]
        )
    return workbook


def excel_safe(value):
    if not isinstance(value, str):
        return value
    if value.startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value
